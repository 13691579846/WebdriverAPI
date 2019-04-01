from openpyxl import load_workbook

class HandleExcel(object):

    def __init__(self):
        self.wb = None
        self.excelPath = None

    def loadWorkBook(self, excelPathName): # 加载excel文件到内存中
        try:
            self.wb = load_workbook(excelPathName)
            self.excelPath = excelPathName
        except Exception as e:
            raise e
        else:
            return self.wb # 返回excel文件对象

    def getSheetName(self):
        try:

            sheetNames = self.wb.sheetnames
        except Exception as e:
            raise e
        else:
            return sheetNames # 获取所有的sheet名称

    def getSheetObj(self, sheetName):
        try:
            sheetObj = self.wb[sheetName]
        except Exception as e:
            raise e
        else:
            return sheetObj # 根据sheet页名称，获取某个sheet页的对象

    def getSheetMaxRow(self, sheetObj):
        try:
            return sheetObj.max_row # 获取某个sheet页的最大数据行
        except Exception as e:
            raise e

    def getSheetMaxColumn(self, sheetObj):
        try:
            return sheetObj.max_column # 获取某个sheet页的最大列数
        except Exception as e:
            raise e

    def getSheetMinRow(self, sheetObj):
        try:
            return sheetObj.min_row # 获取最小行号
        except Exception as e:
            raise e

    def getSheetMinColumn(self, sheetObj):
        try:
            return sheetObj.min_column # 获取最小列号
        except Exception as e:
            raise e

    def getColumnValues(self, sheetObj, colNum):
        # 获取整列数据
        try:
            valueList = []
            for row in range(1, sheetObj.max_row+1):
                values = sheetObj.cell(row, colNum).value
                valueList.append(values)
        except Exception as e:
            raise e
        else:
            return valueList

    def getRowValue(self, sheetObj, rowNum):
        # 获取整行数据
        try:
            valueList = []
            for col in range(1, sheetObj.max_column+1):
                values = sheetObj.cell(rowNum, col).value
                valueList.append(values)
        except Exception as e:
            raise e
        else:
            return valueList

    def getValueOfCell(self, sheetObj, rowNum, colNum):
        # 获取某一个单元格的内容
        try:
            value = sheetObj.cell(rowNum, colNum).value
        except Exception as e:
            raise e
        else:
            return value

    def writeValue(self, sheetObj, rowNum, colNum, value):
        try:
            sheetObj.cell(rowNum, colNum).value = value
            self.wb.save(self.excelPath)
        except Exception as e:
            raise e

if __name__=='__main__':
    p = HandleExcel()
    wb = p.loadWorkBook('test.xlsx')
    sheetNames = p.getSheetName()
    sheetObj = p.getSheetObj(sheetNames[0])
    print('sheetNames =',sheetNames)
    print('最大行为:{},最大列为:{},最小行为:{},最小列为:{}'.format(p.getSheetMaxRow(sheetObj)
                                                   , p.getSheetMaxColumn(sheetObj),
                                                   p.getSheetMinRow(sheetObj),
                                                   p.getSheetMinColumn(sheetObj)))
    print('列数据为：', p.getColumnValues(sheetObj, 1))
    print('行数据为:', p.getRowValue(sheetObj, 1))
    print('1行1列数据为：', p.getValueOfCell(sheetObj, 1,1))
    p.writeValue(sheetObj, 2, 2, 'python')




