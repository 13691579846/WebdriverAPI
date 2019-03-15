from openpyxl import load_workbook

class excelData(object):
    def __init__(self, excelPath, sheetName):
        self.excelPath = excelPath # 初始化文件的路径
        self.sheetName = sheetName # 初始化sheetname

    def read_excel(self):
        wd = load_workbook(self.excelPath) # 打开文件
        sh = wd[self.sheetName] # 获取sheet对象

        dataList = []
        for i in range(2, sh.max_row+1):
            data = []
            data.append(sh.cell(i,2).value)
            data.append(sh.cell(i, 3).value)
            data.append(sh.cell(i, 4).value)
            dataList.append(data)
        return dataList
    def write_excel(self, row, col, value):
        wd = load_workbook(self.excelPath)
        sh = wd[self.sheetName]
        sh.cell(row, col).value = value
        wd.save(self.excelPath)
if __name__=='__main__':
    excel = excelData('./dataexcel.xlsx', 'login')
    print(excel.read_excel())
